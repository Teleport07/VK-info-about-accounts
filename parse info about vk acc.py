import json
import time
import openpyxl
import requests

#подготовить файл excel с id аккаунтов vk
wb = openpyxl.load_workbook('путь к excel файлу')
#на первом листе файла будут id аккаунтов vk
ws = wb['Лист1']
ws1 = wb['Лист2']
part1 = "https://api.vk.com/method/users.get?user_ids="
part_center = "&start_from=&fields=activities, about, bdate, career,city, country, education,followers_count,sex,schools,interests,military,personal,universities"

#в part2 вводим личный vk токен, полученный после инициализации приложения через в vk api
part2 = "vk токен"
paginator = "0"
count = 2

#структура первой страницы - в каждом столбце расположены id аккаунтов для конкретной группы
#поэтому итерация по столбцам - пробегаем по группам
#итерация по строкам - пробегаем по аккаунтам конкретной группы
for col in range(1,4):
    for row in range(2,ws.max_row+1):
        if ws.cell(row = row, column = col).value is not None:
            time.sleep(0.3)
            r = requests.get(part1+str(ws.cell(row = row, column = col).value)+part_center+paginator+part2)
            d = r.json()
            print(d)
            try:
                #верхний предел итерации зависит от количества полей в информации о пользователе, которые мы хотим узнать
                for col1 in range(2,19):
                    if ws1.cell(row = 1, column = col1).value in d['response'][0]:
                        ws1.cell(row = count, column = col1).value = str(d['response'][0][ws1.cell(row = 1, column = col1).value])
            except Exception as ex:
                print('Ошибка')
            count += 1
        else:
            break

#в результате, на втором листе в первом столбце будут id аккаунтов, в других столбцах следующие поля:
#first_name, last_name, activities, about, bdate, career, city, country, education, followers_count, sex,
# schools, interests, military, personal, universities,	career_company, career_city_name, career_from, career_until, career_position

wb.save('путь к excel файлу')